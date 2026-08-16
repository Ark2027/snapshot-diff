"""Load tabular files without caring much what they are.

CSV and Excel both come back as a list of dicts, which is all core.compare
wants. Keeping the loader dumb means the diff logic has no opinion about file
formats and can be tested without touching a disk.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


def load(path: Path | str, sheet: str | None = None) -> list[dict[str, Any]]:
    """Read a CSV or Excel file into a list of row dicts."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel(path, sheet)
    raise ValueError(f"unsupported file type {suffix!r}; expected .csv, .xlsx or .xlsm")


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _load_excel(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading Excel needs openpyxl: pip install 'snapshot-diff[excel]'") from exc

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = _pick_sheet(book, sheet)
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else f"column_{i}" for i, c in enumerate(next(rows))]
        except StopIteration:
            return []
        return [
            dict(zip(header, values))
            for values in rows
            if any(v is not None and str(v).strip() != "" for v in values)
        ]
    finally:
        book.close()


def _pick_sheet(book: Any, sheet: str | None) -> Any:
    """Exact name, then case-insensitive substring, then the first sheet."""
    if sheet is None:
        return book[book.sheetnames[0]]
    if sheet in book.sheetnames:
        return book[sheet]
    for name in book.sheetnames:
        if sheet.lower() in name.lower():
            return book[name]
    raise KeyError(f"no sheet matching {sheet!r}; found {book.sheetnames}")
